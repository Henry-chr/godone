import os
import re
import pydoc
import glob
import pandas
from win32com.client import Dispatch
from openpyxl import load_workbook,Workbook
import os


from pydocx import PyDocX

result = []


def get_all(cwd):
    get_dir = os.listdir(cwd)
    for i in get_dir:
        sub_dir = os.path.join(cwd, i)
        if os.path.isdir(sub_dir):
            get_all(sub_dir)
        else:
            result.append(i)

def del_file():
    path = r'C:\Users\Administrator\Desktop\wdl\WordToExcel'
    print("删除文件：")
    print("-------------------")
    for infile in glob.glob(os.path.join(path, '*.html')):
        os.remove(infile)
        print(infile)

def get_newname(filelist):
    for i in filelist:
        if '-' in i:
            os.rename(i,'现场签证'+str(i))
        else:
            os.rename(i,'变更任务'+str(i))

def xls2xlsx(file_name):
    """
    将xls文件另存为xlsx文件
    :param file_name: 要转换的文件路径
    :returns: new_excel_file_path 返回新的xlsx文件的路径
    """
    excel_file_path = file_name
    import win32com.client
    import os
    # excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    excel = win32com.client.DispatchEx('Excel.Application')
    wb = excel.Workbooks.Open(excel_file_path)
    new_excel_file_path = r"{old_file_path}x".format(old_file_path=excel_file_path.replace('html','xls'))
    if os.path.exists(new_excel_file_path):  # 先删掉新复制的文件
        os.remove(new_excel_file_path)
    wb.SaveAs(new_excel_file_path, FileFormat=51)  # 51 表示的是xlsx格式
    wb.Close()
    excel.Application.Quit()
    # print(new_excel_file_path)

def doc2html(input, output):
    w = Dispatch('Word.Application')
    doc = w.Documents.Open(input, ReadOnly=1)
    doc.SaveAs(output, 8)
    # try:
    #     doc = w.Documents.Open(input, ReadOnly=1)
    #     doc.SaveAs(output, 8)
    #     return True
    # except Exception as e:
    #     return False

def get_html(filelist):
    for i in filelist:
        input = r'C:\Users\Administrator\Desktop\wdl\WordToExcel\\' + str(i)
        output = r'C:\Users\Administrator\Desktop\wdl\WordToExcel\\' + re.sub('docx','html',i)
        print(input)
        print(output)
        doc2html(input,output)


    # for i in filelist:
    #     html = PyDocX.to_html(i)
    #     f = open(re.sub('docx','html',i),'w',encoding='utf-8')
    #     f.write(html)
    #     f.close()

def del_docx(result):
    result_html = []
    for i in result:
        if '.html' in i and '~$' not in i:
            result_html.append(i)
    return result_html

def getxlsx(result):
    result_xlsx = []
    for i in result:
        if '.xlsx' in i and '~$' not in i:
            result_xlsx.append(i)
    return result_xlsx


def to_excel(filelist):
    for i in filelist:
        print(r'C:\Users\Administrator\Desktop\wdl\WordToExcel\\'+str(i))
        xls2xlsx(r'C:\Users\Administrator\Desktop\wdl\WordToExcel\\'+str(i))


def induct(file,newfile,sheets): #文件名 新文件名 sheet列表
	if(not os.path.isfile(file)):
		print(file+" not exists")
	wb=load_workbook(file)
	sheetlist=[]
	for sheet in sheets:
		if(isinstance(sheet,int)): #如果是sheet的编号 换成sheet的名称
			try:
				sheet=wb.sheetnames[sheet]
			except Exception as e:
				print('页序号不存在')
				continue
		sheetlist.append(sheet)
	for sheetname in wb.sheetnames:
		if(sheetname not in sheetlist):
			ws=wb[sheetname]
			wb.remove(ws)
	wb.save(newfile)								#保存到newfile

def connect(result):
    for i in result:
        file = r'C:\Users\Administrator\Desktop\wdl\Excel\\'+str(i)
        newfile = r'C:\Users\Administrator\Desktop\wdl\output.xlsx'
        sheets = []
        sheets.append(i.replace('.xlsx',''))
        induct(file,newfile,sheets)
        print(file+'合并完成！！！！')

def get_docx(filelist):
    list_html1 = []
    for i in filelist:
        list_html2 = []
        with open(i,'rb') as f:
            df = pandas.read_html(f.read())
            list_html2.append(i.replace('.html',''))
            list_html2.append(df[0])
        list_html1.append(list_html2)
    print(list_html1)
    with pandas.ExcelWriter(r'C:\Users\Administrator\Desktop\wdl\output.xlsx') as bb:
        for j in range(len(list_html1)):
            a = list_html1[j][0]
            b = list_html1[j][1]
            b.to_excel(bb,sheet_name=a)

if __name__ == "__main__":
    os.chdir(r'C:\Users\Administrator\Desktop\wdl\Excel')
    # del_file()
    get_all(r'C:\Users\Administrator\Desktop\wdl\Excel')
    print(result)
    # get_newname(result)
    # get_html(result)
    # result1 = del_docx(result)
    # print(result1)
    # to_excel(result1)
    # result2 = getxlsx(result)
    # print(result2)
    connect(result)