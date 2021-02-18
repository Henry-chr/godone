from openpyxl import load_workbook,Workbook
import os
def induct(file,newfile,sheets): #文件名 新文件名 sheet列表
	if(not os.path.isfile(file)):
		print(file+" not exists")
	wb=load_workbook(file)
	sheetlist=[]
	for sheet in sheets:
		if(isinstance(sheet,int)): #如果是sheet的编号 换成sheet的名称
			try:
				sheet=wb.sheetnames[sheet]
			except Exception as e:
				print('页序号不存在')
				continue
		sheetlist.append(sheet)
	for sheetname in wb.sheetnames:
		if(sheetname not in sheetlist):
			ws=wb[sheetname]
			wb.remove(ws)
	wb.save(newfile)								#保存到newfile

induct(r'C:\Users\Administrator\Desktop\wdl\Excel\变更任务01.xlsx',r'C:\Users\Administrator\Desktop\wdl\Excel\test.xlsx',['变更任务01'])
